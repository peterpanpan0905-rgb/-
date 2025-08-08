import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import openpyxl
import os
import shutil
from datetime import datetime
from openpyxl.styles import PatternFill, Font
import sys


class AttendanceProcessorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("考勤表导出整理工具")
        self.root.geometry("800x580")
        self.root.configure(bg="#f0f2f5")

        # 设置应用图标
        self.set_app_icon()

        # 创建样式
        self.create_styles()

        # 创建主框架
        self.main_frame = ttk.Frame(root, style="Main.TFrame")
        self.main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)

        # 标题区域
        self.header_frame = ttk.Frame(self.main_frame)
        self.header_frame.pack(fill=tk.X, pady=(0, 20))

        ttk.Label(self.header_frame, text="考勤表导出整理工具",
                  style="Header.TLabel").pack(side=tk.LEFT)

        # 文件选择区域
        self.create_file_selection()

        # 应出勤天数输入框
        self.create_working_days_input()
        #输入当月月份
        self.create_month_input()
        # 处理按钮
        self.btn_frame = ttk.Frame(self.main_frame)
        self.btn_frame.pack(fill=tk.X, pady=20)

        btn_container = ttk.Frame(self.btn_frame)
        btn_container.pack(expand=True)

        self.process_btn = ttk.Button(btn_container, text="生成考勤表",
                                      command=self.process_files,
                                      style="Primary.TButton", width=20)
        self.process_btn.pack(pady=10)

        # 状态区域
        self.status_frame = ttk.Frame(self.main_frame, style="Status.TFrame")
        self.status_frame.pack(fill=tk.X, pady=(5, 0))

        self.status_text = tk.StringVar(value="准备就绪")
        status_label = ttk.Label(self.status_frame, textvariable=self.status_text,
                                 style="Status.TLabel")
        status_label.pack(side=tk.LEFT, padx=10, pady=5)

        # 底部信息
        self.footer_frame = ttk.Frame(self.main_frame, style="Footer.TFrame")
        self.footer_frame.pack(fill=tk.X, pady=(15, 0))

        ttk.Label(self.footer_frame, text="© 2025 考勤表导出整理工具 | 版本 1.0",
                  style="Footer.TLabel").pack(side=tk.RIGHT, padx=10, pady=5)

    def set_app_icon(self):
        """设置应用图标"""
        try:
            if sys.platform == "win32":
                self.root.iconbitmap("icon.ico")  # Windows
            # 其他平台可以添加相应处理
        except:
            pass

    def create_styles(self):
        """创建应用程序样式"""
        style = ttk.Style()
        style.theme_use("clam")

        # 主框架样式
        style.configure("Main.TFrame", background="#ffffff", borderwidth=0)

        # 标题样式
        style.configure("Header.TLabel",
                        font=("Microsoft YaHei", 16, "bold"),
                        background="#ffffff",
                        foreground="#2c3e50")

        # 卡片样式
        style.configure("Card.TLabelframe",
                        borderwidth=1,
                        relief="solid",
                        bordercolor="#e1e4e8",
                        background="#ffffff",
                        foreground="#2c3e50",
                        font=("Microsoft YaHei", 10, "bold"))

        style.configure("Card.TLabelframe.Label",
                        background="#ffffff",
                        foreground="#2c3e50",
                        font=("Microsoft YaHei", 10, "bold"))

        # 输入框样式
        style.configure("Custom.TEntry",
                        borderwidth=1,
                        relief="solid",
                        bordercolor="#ced4da",
                        padding=5,
                        font=("Microsoft YaHei", 10))

        # 按钮样式
        style.configure("Accent.TButton",
                        font=("Microsoft YaHei", 10),
                        foreground="#ffffff",
                        background="#6c757d",
                        borderwidth=0,
                        padding=4)

        style.map("Accent.TButton",
                  background=[('active', '#5a6268'), ('pressed', '#545b62')])

        style.configure("Primary.TButton",
                        font=("Microsoft YaHei", 10, "bold"),
                        foreground="#ffffff",
                        background="#007bff",
                        borderwidth=0,
                        padding=8)

        style.map("Primary.TButton",
                  background=[('active', '#0069d9'), ('pressed', '#0062cc')])

        # 标签样式
        style.configure("File.TLabel",
                        font=("Microsoft YaHei", 10),
                        background="#ffffff",
                        foreground="#495057")

        # 参数标签样式
        style.configure("Param.TLabel",
                        font=("Microsoft YaHei", 10),
                        background="#ffffff",
                        foreground="#495057")

        # 状态栏样式
        style.configure("Status.TFrame",
                        background="#e9ecef",
                        borderwidth=1,
                        relief="solid",
                        bordercolor="#dee2e6")

        style.configure("Status.TLabel",
                        font=("Microsoft YaHei", 9),
                        background="#e9ecef",
                        foreground="#6c757d")

        # 底部样式
        style.configure("Footer.TFrame",
                        background="#f8f9fa",
                        borderwidth=1,
                        relief="solid",
                        bordercolor="#dee2e6")

        style.configure("Footer.TLabel",
                        font=("Microsoft YaHei", 8),
                        background="#f8f9fa",
                        foreground="#6c757d")

    def create_file_selection(self):
        """创建文件选择区域"""
        # 月度请假单据选择
        jt_frame = ttk.LabelFrame(self.main_frame, text=" 1. 选择钉钉导出月度请假单据原始数据 ",
                                  style="Card.TLabelframe")
        jt_frame.pack(fill=tk.X, pady=5)

        ttk.Label(jt_frame, text="月度请假单据:",
                  style="File.TLabel").grid(row=0, column=0, padx=(10, 5), pady=5, sticky="e")

        self.jt_entry = ttk.Entry(jt_frame, width=50, style="Custom.TEntry")
        self.jt_entry.grid(row=0, column=1, padx=5, pady=5, sticky="ew")

        ttk.Button(jt_frame, text="浏览...",
                   command=self.select_jt_file,
                   style="Accent.TButton").grid(row=0, column=2, padx=5, pady=5)

        # 日考勤汇总表选择
        jsk_frame = ttk.LabelFrame(self.main_frame, text=" 2. 选择钉钉导出统计当月的日考勤汇总表 ",
                                   style="Card.TLabelframe")
        jsk_frame.pack(fill=tk.X, pady=5)

        ttk.Label(jsk_frame, text="日考勤汇总表:",
                  style="File.TLabel").grid(row=0, column=0, padx=(10, 5), pady=5, sticky="e")

        self.jsk_entry = ttk.Entry(jsk_frame, width=50, style="Custom.TEntry")
        self.jsk_entry.grid(row=0, column=1, padx=5, pady=5, sticky="ew")

        ttk.Button(jsk_frame, text="浏览...",
                   command=self.select_jsk_file,
                   style="Accent.TButton").grid(row=0, column=2, padx=5, pady=5)

    def create_working_days_input(self):
        """创建应出勤天数输入区域"""
        param_frame = ttk.LabelFrame(self.main_frame, text=" 3. 设置考勤参数 ",
                                     style="Card.TLabelframe")
        param_frame.pack(fill=tk.X, pady=10)

        # 应出勤天数输入
        work_days_frame = ttk.Frame(param_frame)
        work_days_frame.pack(fill=tk.X, padx=10, pady=5)

        ttk.Label(work_days_frame, text="应出勤天数:",
                  style="Param.TLabel").pack(side=tk.LEFT, padx=(0, 10))

        self.working_days = tk.StringVar(value="22")
        work_days_entry = ttk.Entry(work_days_frame, textvariable=self.working_days,
                                    width=10, style="Custom.TEntry")
        work_days_entry.pack(side=tk.LEFT)

        # 添加说明标签
        # ttk.Label(param_frame, text="注：程序使用内置模板文件，无需选择",
        #           style="Param.TLabel", foreground="#6c757d").pack(side=tk.BOTTOM, padx=10, pady=(0, 5))

    def create_month_input(self):
        """创建月份输入区域"""
        param_frame = ttk.LabelFrame(self.main_frame, text=" 4. 月份 ",
                                     style="Card.TLabelframe")
        param_frame.pack(fill=tk.X, pady=5)

        # 月份天数输入
        month_days_frame = ttk.Frame(param_frame)
        month_days_frame.pack(fill=tk.X, padx=10, pady=5)

        ttk.Label(month_days_frame, text="月份:",
                  style="Param.TLabel").pack(side=tk.LEFT, padx=(0, 10))

        self.month_days = tk.StringVar(value="")
        month_days_entry = ttk.Entry(month_days_frame, textvariable=self.month_days,
                                    width=10, style="Custom.TEntry")
        month_days_entry.pack(side=tk.LEFT)

        # 添加说明标签
        # ttk.Label(param_frame, text="注：程序使用内置模板文件，无需选择",
        #           style="Param.TLabel", foreground="#6c757d").pack(side=tk.BOTTOM, padx=10, pady=(0, 5))

    def select_jt_file(self):
        file_path = filedialog.askopenfilename(title="选择月度请假单据", filetypes=[("Excel文件", "*.xlsx")])
        if file_path:
            self.jt_entry.delete(0, tk.END)
            self.jt_entry.insert(0, file_path)
            self.status_text.set(f"已选择月度请假单据: {os.path.basename(file_path)}")

    def select_jsk_file(self):
        file_path = filedialog.askopenfilename(title="选择日考勤汇总表", filetypes=[("Excel文件", "*.xlsx")])
        if file_path:
            self.jsk_entry.delete(0, tk.END)
            self.jsk_entry.insert(0, file_path)
            self.status_text.set(f"已选择日考勤汇总表: {os.path.basename(file_path)}")

    def get_template_path(self, template_name):
        """获取内置模板文件路径"""
        # 尝试获取当前脚本所在目录
        if getattr(sys, 'frozen', False):
            # 打包后的可执行文件路径
            base_path = os.path.dirname(sys.executable)
        else:
            # 脚本文件路径
            base_path = os.path.dirname(os.path.abspath(__file__))

        # 构建模板路径
        template_path = os.path.join(base_path, "templates", template_name)

        # 检查模板文件是否存在
        if not os.path.exists(template_path):
            messagebox.showerror("错误", f"找不到内置模板文件: {template_name}\n请确保程序目录下存在templates文件夹")
            return None

        return template_path

    def process_files(self):
        # 获取文件路径
        leave_path = self.jt_entry.get()
        day_re_path = self.jsk_entry.get()

        # 获取应出勤天数
        working_days = self.working_days.get()

        # 月份获取
        month_day = self.month_days.get()

        # 验证文件是否选择
        if not all([leave_path, day_re_path]):
            messagebox.showerror("错误", "请选择所有必要的文件")
            return

        # 验证应出勤天数
        if not working_days.isdigit() or int(working_days) <= 0:
            messagebox.showerror("错误", "应出勤天数必须是大于0的整数")
            return

        working_days = int(working_days)

        try:
            # 更新状态
            self.status_text.set("处理中，请稍候...")
            self.root.update_idletasks()

            # 获取内置模板路径
            template_path = self.get_template_path("考勤表模板.xlsx")
            leave_template_path = self.get_template_path("请假表模板.xlsx")

            if not template_path or not leave_template_path:
                return

            # 创建输出文件名
            now = datetime.now()
            year_month_day = now.strftime("%Y%m%d")
            output_dir = os.path.dirname(day_re_path)  # 使用日考勤汇总表所在目录
            output_file = os.path.join(output_dir, f"考勤表{year_month_day}.xlsx")

            # 请假表处理
            self.leave_deal(leave_path, leave_template_path, 'leave_deal_temp.xlsx',month_day)

            # 复制模板文件
            shutil.copy(template_path, output_file)
            wb = openpyxl.load_workbook(day_re_path, data_only=False)
            wn = openpyxl.load_workbook(output_file, data_only=False)  # 使用输出文件

            # 处理考勤表1
            n_data = self.process_1_sheet(wb, wn, output_file, working_days, month_day)

            # 处理考勤表2
            self.process_2_sheet('leave_deal_temp.xlsx', wn, n_data, output_file, month_day)

            # 清理临时文件
            if os.path.exists('leave_deal_temp.xlsx'):
                os.remove('leave_deal_temp.xlsx')

            messagebox.showinfo("成功", f"考勤表生成成功!\n文件路径: {output_file}")
            self.status_text.set(f"考勤表生成成功: {os.path.basename(output_file)}")

        except Exception as e:
            messagebox.showerror("错误", f"处理过程中发生错误: {str(e)}")
            self.status_text.set(f"处理失败: {str(e)}")

    def process_1_sheet(self, wb, wn, output_path, working_days, month_day):
        # 获取sheet1
        profit_sheet = wb[wb.sheetnames[0]]
        source_data = {}
        chidao_time = {}
        zaotui_time = {}
        shangban_forget = {}
        xiaban_forget = {}
        over_time_total = {}
        chuqin_time = {}
        n_data = {}
        na_data = {}
        code = 2
        item = 1
        # 构建唯一码到行索引的映射
        for row in profit_sheet.iter_rows(min_row=5, values_only=True):
            code += 1
            # 取科目信息 (A-D列)
            a_val = row[0]  # 姓名
            b_val = row[2]  # 部门
            c_val = row[4]  # 职位
            d_val = row[6]  # 日期
            e_val = row[8]  # 班次
            f_val = row[9]  # 上班打卡时间
            g_val = row[10]  # 上班打卡结果
            h_val = row[11]  # 下班打卡时间
            i_val = row[12]  # 下班打卡结果
            j_val = row[24]  # 工作时长
            k_val = row[26]  # 迟到时长
            l_val = row[31]  # 早退时长
            m_val = row[32]  # 上班缺卡次数
            o_val = row[33]  # 下班缺卡次数
            p_val = row[22]  # 出勤天数
            if a_val == '肖蝶':
                pass
            else:
                if not j_val:
                    over_time_val = 0
                else:
                    over_time_val = (float(j_val) - 450) / 60

                source_data[code] = (a_val, b_val, c_val, d_val, e_val, f_val, g_val, h_val, i_val, j_val, k_val, l_val,
                                     m_val, o_val, over_time_val)

                # 迟到早退时长上下班打卡次数统计
                if k_val:
                    if a_val in chidao_time:
                        chidao_time[a_val] += float(k_val)
                    else:
                        chidao_time[a_val] = float(k_val)

                if l_val:
                    if a_val in zaotui_time:
                        zaotui_time[a_val] += float(l_val)
                    else:
                        zaotui_time[a_val] = float(l_val)

                if m_val:
                    if a_val in shangban_forget:
                        shangban_forget[a_val] += float(m_val)
                    else:
                        shangban_forget[a_val] = float(m_val)

                if o_val:
                    if a_val in xiaban_forget:
                        xiaban_forget[a_val] += float(o_val)
                    else:
                        xiaban_forget[a_val] = float(o_val)
                if p_val:
                    if a_val in chuqin_time:
                        chuqin_time[a_val] += float(p_val)
                    else:
                        chuqin_time[a_val] = float(p_val)

                if a_val in over_time_total:
                    over_time_total[a_val] += float(over_time_val)
                else:
                    over_time_total[a_val] = float(over_time_val)

                con_val = (a_val, b_val)
                # if con_val not in n_data and a_val[-1] != "）":
                if con_val not in n_data:
                    n_data[con_val] = item
                    if a_val not in na_data:
                        na_data[a_val] = item
                    item += 1

        # 处理考勤表1数据
        wn_sheet = wn[wn.sheetnames[0]]
        wd_sheet = wn[wn.sheetnames[1]]

        # 月份天数计算
        if month_day in ['1', '3', '5', '7', '8', '10', '12']:
            day = 30
        elif month_day in ['4', '6', '9', '10']:
            day = 29
        else:
            try:
                wn_sheet.merge_cells(start_row=3, start_column=1, end_row=3 + 28, end_column=1)
                day = 28
                wn_sheet.unmerge_cells(start_row=3, start_column=1, end_row=3 + 28, end_column=1)
            except:
                day = 27

        for row in range(day+4, len(source_data)):
            a_val, b_val, c_val, d_val, e_val, f_val, g_val, h_val, i_val, j_val, k_val, l_val, m_val, o_val, over_time_val = \
            source_data[row]
            row = row - day - 1
            wn_sheet.cell(row=row, column=1,value = a_val)
            wn_sheet.cell(row=row, column=2,value = b_val)
            wn_sheet.cell(row=row, column=3,value = c_val)
            wn_sheet.cell(row=row, column=4,value = d_val)
            wn_sheet.cell(row=row, column=5,value = e_val)
            wn_sheet.cell(row=row, column=6,value = f_val)
            wn_sheet.cell(row=row, column=7,value = g_val)
            wn_sheet.cell(row=row, column=8,value = h_val)
            wn_sheet.cell(row=row, column=9,value = i_val)
            wn_sheet.cell(row=row, column=10).value = chidao_time[a_val] if a_val in chidao_time else 0
            wn_sheet.cell(row=row, column=11).value = zaotui_time[a_val] if a_val in zaotui_time else 0
            wn_sheet.cell(row=row, column=12).value = shangban_forget[a_val] if a_val in shangban_forget else 0
            wn_sheet.cell(row=row, column=13).value = xiaban_forget[a_val] if a_val in xiaban_forget else 0
            wn_sheet.cell(row=row, column=14).value = over_time_total[a_val]

        # 合并单元格
        for mer in range(3, len(source_data), day+1):
            wn_sheet.merge_cells(start_row=mer, start_column=1, end_row=mer + day, end_column=1)
            wn_sheet.merge_cells(start_row=mer, start_column=10, end_row=mer + day, end_column=10)
            wn_sheet.merge_cells(start_row=mer, start_column=11, end_row=mer + day, end_column=11)
            wn_sheet.merge_cells(start_row=mer, start_column=12, end_row=mer + day, end_column=12)
            wn_sheet.merge_cells(start_row=mer, start_column=13, end_row=mer + day, end_column=13)
            wn_sheet.merge_cells(start_row=mer, start_column=14, end_row=mer + day, end_column=14)

        # 填充颜色
        chidao = PatternFill("solid", fgColor="CCFFCC")
        waichu = PatternFill("solid", fgColor="FFCC99")
        queka = PatternFill("solid", fgColor="FF8080")
        qingjia = PatternFill("solid", fgColor="FFC000")
        waiqing = PatternFill("solid", fgColor="FFFF00")
        chuchai = PatternFill("solid", fgColor="00B0F0")
        for row in wn_sheet.iter_rows(min_row=3, min_col=4, max_col=9, values_only=False):
            for cell in row:
                if cell.value:
                    if cell.value == "请假":
                        cell.fill = qingjia
                    elif cell.value == "出差":
                        cell.fill = chuchai
                    elif cell.value == "外勤":
                        cell.fill = waiqing
                    elif cell.value == "迟到":
                        cell.fill = chidao
                    elif cell.value == "外出" or cell.value == "补卡审批通过":
                        cell.fill = waichu
                    elif cell.value == "缺卡":
                        cell.fill = queka
                    if isinstance(cell.value, str):
                        if cell.value.endswith("六") or cell.value.endswith("日"):
                            cell.font = Font(name="新宋体", size=12, color="FF0000", bold=True)

        # 处理考勤表2
        for i in n_data.keys():
            a_val, b_val = i
            row = n_data[i] + 4
            wd_sheet.cell(row=row, column=1).value = n_data[i]
            wd_sheet.cell(row=row, column=2).value = a_val
            wd_sheet.cell(row=row, column=3).value = b_val
            wd_sheet.cell(row=row, column=6).value = working_days  # 使用传入的应出勤天数
            wd_sheet.cell(row=row, column=7).value = chuqin_time[a_val] if a_val in chuqin_time else 0
            wd_sheet.cell(row=row, column=8).value = '=G' + f'{row}' + '-I' + f'{row}'
            wd_sheet.cell(row=row, column=23).value = over_time_total[a_val]
            if a_val in chidao_time:
                wd_sheet.cell(row=row, column=18).value = chidao_time[a_val]
            if a_val in zaotui_time:
                wd_sheet.cell(row=row, column=19).value = zaotui_time[a_val]
            if a_val in shangban_forget:
                wd_sheet.cell(row=row, column=21).value = shangban_forget[a_val]
            if a_val in xiaban_forget:
                wd_sheet.cell(row=row, column=22).value = xiaban_forget[a_val]

        # 保存结果
        wn.save(output_path)
        return na_data

    def process_2_sheet(self, leave_path, wn, na_data, output_path,month_day):
        wb = openpyxl.load_workbook(leave_path, data_only=True)
        wb_sheet = wb[wb.sheetnames[0]]


        data_leave = {}
        # 请假类型与时间绑定到字典
        for row in wb_sheet.iter_rows(min_row=2, values_only=True):
            name_val = row[0]
            apart = row[1]
            leave_stype = row[2]
            f = (name_val, apart, leave_stype)
            if f not in data_leave:
                temp = float(row[6]) / 7.5
                data_leave[f] = temp

        # 获取考勤表2的工作表
        ws_sheet = wn[wn.sheetnames[1]]
        ws_sheet['A1'].value = 'jointelli-0'+f'{month_day}'+'月考勤表'
        ws_sheet['A1'].font = Font(name="微软雅黑", size=16, color="000000", bold=True)
        # 请假信息填到考勤表2
        for i in data_leave.keys():
            a, b, c = i
            if a in na_data:
                n_row = na_data[a] + 4
            else:
                print(f"未找到员工: {a}")
                continue  # 跳过未找到的员工

            if c == "事假":
                ws_sheet.cell(row=n_row, column=9).value = data_leave[(a, b, c)]
            elif c == "调休":
                ws_sheet.cell(row=n_row, column=10).value = data_leave[(a, b, c)]
            elif c == "年假":
                ws_sheet.cell(row=n_row, column=11).value = data_leave[(a, b, c)]
            elif c == "病假":
                ws_sheet.cell(row=n_row, column=12).value = data_leave[(a, b, c)]
            elif c == "婚假":
                ws_sheet.cell(row=n_row, column=13).value = data_leave[(a, b, c)]
            elif c == "陪产假":
                ws_sheet.cell(row=n_row, column=14).value = data_leave[(a, b, c)]
            elif c == "产前假":
                ws_sheet.cell(row=n_row, column=15).value = data_leave[(a, b, c)]
            elif c == "产假":
                ws_sheet.cell(row=n_row, column=16).value = data_leave[(a, b, c)]
            elif c == "产检假":
                ws_sheet.cell(row=n_row, column=17).value = data_leave[(a, b, c)]

        # 保存结果
        wn.save(output_path)

    def leave_deal(self, leave_file, leave_moban, output_path,month_day):
        leave_path = leave_file
        wb = openpyxl.load_workbook(leave_path, data_only=True)
        sheet1 = wb[wb.sheetnames[0]]
        data_dir = {}
        row_data = 2
        shijia_time_sum = {}
        tiaoxiu_time_sum = {}
        nianjia_time_sum = {}
        bingjia_time_sum = {}
        hunjia_time_sum = {}
        peichan_time_sum = {}
        chanqian_time_sum = {}
        chanjia_time_sum = {}
        chanjian_time_sum = {}
        search_text = "(已离职)"
        # 添加行号索引
        for row_idx, row in enumerate(sheet1.iter_rows(min_row=2, min_col=10, max_col=10, values_only=True), start=2):
            for cell in row:
                # 添加None值检查
                if cell is not None and isinstance(cell, str) and cell.endswith(search_text):
                    ff = f'{cell[:-5]}（离职）'  # 使用全角括号保持一致性
                    # 使用行号索引(row_idx)而不是row元组
                    sheet1.cell(row=row_idx, column=10).value = ff

        for row in sheet1.iter_rows(min_row=2, max_row=sheet1.max_row, values_only=True):
            #检查是否为当月休假
            d_val = row[15]  # 开始时间
            if d_val[6] == f'{month_day}':

                a_val = row[9]  # 请假发起人
                b_val = row[10]  # 部门
                c_val = row[14]  # 请假类型

                e_val = row[16]  # 结束时间
                f_val = row[17]  # 时长
                g_val = row[2]  # 完成状态
                h_val = row[3]  # 审批状态


                # 转换时长格式
                if f_val and isinstance(f_val, str):
                    if f_val.endswith("时"):
                        f_val = float(f_val[:-2])
                    elif f_val.endswith("天"):
                        f_val = float(f_val[:-1]) * 7.5
                else:
                    f_val = 0.0  # 默认值

                # 只处理审批通过且完成的请假记录
                if g_val == "完成" and h_val == "同意":
                    data_dir[row_data] = (a_val, b_val, c_val, d_val, e_val, f_val)
                    row_data += 1

                # 统计总时长
                if c_val == "事假":
                    if a_val in shijia_time_sum:
                        shijia_time_sum[a_val] += f_val
                    else:
                        shijia_time_sum[a_val] = f_val
                elif c_val == "调休":
                    if a_val in tiaoxiu_time_sum:
                        tiaoxiu_time_sum[a_val] += f_val
                    else:
                        tiaoxiu_time_sum[a_val] = f_val
                elif c_val == "年假":
                    if a_val in nianjia_time_sum:
                        nianjia_time_sum[a_val] += f_val
                    else:
                        nianjia_time_sum[a_val] = f_val
                elif c_val == "病假":
                    if a_val in bingjia_time_sum:
                        bingjia_time_sum[a_val] += f_val
                    else:
                        bingjia_time_sum[a_val] = f_val
                elif c_val == "婚假":
                    if a_val in hunjia_time_sum:
                        hunjia_time_sum[a_val] += f_val
                    else:
                        hunjia_time_sum[a_val] = f_val
                elif c_val == "陪产假":
                    if a_val in peichan_time_sum:
                        peichan_time_sum[a_val] += f_val
                    else:
                        peichan_time_sum[a_val] = f_val
                elif c_val == "产前假":
                    if a_val in chanqian_time_sum:
                        chanqian_time_sum[a_val] += f_val
                    else:
                        chanqian_time_sum[a_val] = f_val
                elif c_val == "产假":
                    if a_val in chanjia_time_sum:
                        chanjia_time_sum[a_val] += f_val
                    else:
                        chanjia_time_sum[a_val] = f_val
                elif c_val == "产检假":
                    if a_val in chanjian_time_sum:
                        chanjian_time_sum[a_val] += f_val
                    else:
                        chanjian_time_sum[a_val] = f_val

        # 填表
        ws = openpyxl.load_workbook(leave_moban, data_only=True)
        sheet2 = ws[ws.sheetnames[0]]

        for row in range(2, len(data_dir) + 1):  # 修正循环范围
            if row not in data_dir:
                continue

            a_val, b_val, c_val, d_val, e_val, f_val = data_dir[row]

            sheet2.cell(row=row, column=1, value=a_val)
            sheet2.cell(row=row, column=2, value=b_val)
            sheet2.cell(row=row, column=3, value=c_val)
            sheet2.cell(row=row, column=4, value=d_val)
            sheet2.cell(row=row, column=5, value=e_val)
            sheet2.cell(row=row, column=6, value=f_val)

            # 设置总时长
            if c_val == "事假":
                sheet2.cell(row=row, column=7, value=shijia_time_sum.get(a_val, 0))
            elif c_val == "调休":
                sheet2.cell(row=row, column=7, value=tiaoxiu_time_sum.get(a_val, 0))
            elif c_val == "年假":
                sheet2.cell(row=row, column=7, value=nianjia_time_sum.get(a_val, 0))
            elif c_val == "病假":
                sheet2.cell(row=row, column=7, value=bingjia_time_sum.get(a_val, 0))
            elif c_val == "婚假":
                sheet2.cell(row=row, column=7, value=hunjia_time_sum.get(a_val, 0))
            elif c_val == "陪产假":
                sheet2.cell(row=row, column=7, value=peichan_time_sum.get(a_val, 0))
            elif c_val == "产前假":
                sheet2.cell(row=row, column=7, value=chanqian_time_sum.get(a_val, 0))
            elif c_val == "产假":
                sheet2.cell(row=row, column=7, value=chanjia_time_sum.get(a_val, 0))
            elif c_val == "产检假":
                sheet2.cell(row=row, column=7, value=chanjian_time_sum.get(a_val, 0))

            # 设置天数列
            sheet2.cell(row=row, column=8).value = f'=G{row}/7.5'
            sheet2.cell(row=row, column=8).number_format = '0.00'

        # 保存结果
        ws.save(output_path)
        ws.close()


# 运行应用程序
if __name__ == "__main__":
    root = tk.Tk()
    app = AttendanceProcessorApp(root)
    root.mainloop()